"""
Report helper for DayFlow HRMS.

Generates:
  - Salary slip PDFs (single employee, single month)
  - Attendance / Leave / Payroll / Employee report PDFs
  - Excel exports for reports
"""

import io
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SAGE = colors.HexColor("#6BBF8A")
SAGE_DARK = colors.HexColor("#5AA978")
TEXT_DARK = colors.HexColor("#24332A")
BORDER = colors.HexColor("#DDE8DF")

MONTH_NAMES = [
    "", "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]


def _styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="DFTitle", fontSize=18, textColor=SAGE_DARK, spaceAfter=4, leading=22))
    styles.add(ParagraphStyle(name="DFSubtitle", fontSize=10, textColor=TEXT_DARK, spaceAfter=10))
    styles.add(ParagraphStyle(name="DFSection", fontSize=12, textColor=SAGE_DARK, spaceBefore=14, spaceAfter=6))
    return styles


def generate_salary_slip_pdf(employee, month, year, company_name="DayFlow Technologies Pvt. Ltd."):
    """Return a BytesIO buffer containing a salary slip PDF for one employee/month."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = _styles()
    elements = []

    elements.append(Paragraph(company_name, styles["DFTitle"]))
    elements.append(Paragraph(f"Salary Slip — {MONTH_NAMES[month]} {year}", styles["DFSubtitle"]))

    basic = employee["basic_salary"]
    allowance = employee["allowance"]
    deduction = employee["deduction"]
    net = basic + allowance - deduction

    info_data = [
        ["Employee Name", employee["full_name"], "Employee ID", employee["employee_code"]],
        ["Department", employee["department"] or "-", "Job Title", employee["job_title"] or "-"],
        ["Pay Period", f"{MONTH_NAMES[month]} {year}", "Generated On", datetime.now().strftime("%d %b %Y")],
    ]
    info_table = Table(info_data, colWidths=[35 * mm, 55 * mm, 35 * mm, 45 * mm])
    info_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (-1, -1), TEXT_DARK),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("Earnings & Deductions", styles["DFSection"]))
    salary_data = [
        ["Component", "Amount (Rs.)"],
        ["Basic Pay", f"{basic:,.2f}"],
        ["Allowance", f"{allowance:,.2f}"],
        ["Gross Earnings", f"{basic + allowance:,.2f}"],
        ["Deduction", f"{deduction:,.2f}"],
        ["Net Salary", f"{net:,.2f}"],
    ]
    salary_table = Table(salary_data, colWidths=[85 * mm, 85 * mm])
    salary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), SAGE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F4FBF6")),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
    ]))
    elements.append(salary_table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "This is a system-generated salary slip and does not require a signature.",
        styles["DFSubtitle"]
    ))

    doc.build(elements)
    buf.seek(0)
    return buf


def generate_report_pdf(title, headers, rows, subtitle=""):
    """Generic tabular report PDF (attendance / leave / payroll / employee)."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = _styles()
    elements = [Paragraph(title, styles["DFTitle"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["DFSubtitle"]))
    elements.append(Spacer(1, 8))

    data = [headers] + rows
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), SAGE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4FBF6")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf


def generate_report_excel(title, headers, rows):
    """Generic tabular report as an Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Report"

    ws.append(headers)
    header_fill = PatternFill(start_color="6BBF8A", end_color="6BBF8A", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(12, min(40, max_len + 4))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
